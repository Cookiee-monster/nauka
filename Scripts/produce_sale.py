#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet

import os, openpyxl, time
from openpyxl.styles import Font, NamedStyle

start = time.time()

os.chdir(r"C:\Pajtek")

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

price_updates = {"Garlic": 3.07,
                 "Celery": 1.19,
                 "Lemon": 1.27}

print("Workbook was opened correctly")
bold14font = Font(size = 14, bold=True, color = "ff0000")


# Loop for all rows and check whether there is anything to correct

for row_num in range(2, sheet.max_row):
    # skip the first row
    produce_name = sheet.cell(row = row_num, column = 1).value
    if produce_name in price_updates:
        sheet.cell(row = row_num, column = 2).value = price_updates[produce_name]
        sheet.cell(row = row_num, column = 1).font = bold14font

sheet.row_dimensions[1].height = 110
sheet.column_dimensions['B'].width = 20
sheet.merge_cells("A1:D12")
sheet.freeze_panes = "E12"


wb.save("updatedproducesales.xlsx")

end = time.time()

elapsed = end - start

print("Script has finished in {} seconds".format(round(elapsed)))